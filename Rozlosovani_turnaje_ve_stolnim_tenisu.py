import pandas as pd
import random
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import Workbook

class Osoba:
    def __init__(self, jmeno, prijmeni):
        self.jmeno = jmeno
        self.prijmeni = prijmeni
    def __repr__(self):
        return f"Osoba ({self.jmeno}, {self.prijmeni})"

class Hrac(Osoba): #Hráč dědí z Osoby, reprezentuje každého hráče turnaje
    def __init__(self, jmeno, prijmeni, gender, nasazeni, klub):
        super().__init__(jmeno, prijmeni)
        self.gender = gender
        self.klub = klub
        self.nasazeni = nasazeni
        self.kos = None
    def __repr__(self):
        return f"Hrac ({self.jmeno}, {self.prijmeni}, {self.klub})"
    
    def nacist_hrace(soubor): #Python přečte Excel a přiradí k jednotlivým proměnným specifikace hráčů dle Excelu
        df = pd.read_excel(soubor, header=1)
        hraci = []
        for _, row in df.iterrows():
            hrac = Hrac(jmeno = row["Jméno"],
                prijmeni = row["Přijmení"],
                gender= row["Gender"],
                klub = row["Klub"],
                nasazeni = row["Nasazení"],
                )
            hraci.append(hrac)
        return hraci

    def rozradit_gender(hraci): #Metoda, která rozřadí Hráče pomocí genderu
        muzi = []
        zeny = []

        #Rozřazení probíhá podle toho, co je napsané v kolonce Gender u každého hráče
        for hrac in hraci: 
            if hrac.gender.lower() in ["m", "muž"]: #U každého genderu více možností pro univerzálnost kódu
                muzi.append(hrac)
            elif hrac.gender.lower() in ["f", "ž", "žena"]:#Zároveň gender převeden na malá písmena, znovu pro zvýšení univerzálnosti
                zeny.append(hrac)
        return muzi, zeny
    
    def serazeni_a_prepis_nasazeni(hraci): #Tato metoda seřadí hráče podle nasazení a přepíše jim globální nasazení ze žebříčku do lokálního nasazení turnaje
        serazeni_hraci = sorted(hraci, key=lambda hrac: hrac.nasazeni) #seřazení hráčů
        for index, hrac in enumerate(serazeni_hraci, start = 1): #přepsání nasazení
            hrac.nasazeni = index
        return serazeni_hraci
    
    def rozdeleni_na_kose(serazeni_hraci, pocet_skupin):#Metoda, která rozdělí hráče do podskupinek dle jeho nasazení a počtu skupin
        delic = pocet_skupin # Proměnná, která rozděluje hráče na podskupinky a dělí tyto podskupinky (12. hráč je jednička, 13. dvojka) 
    #př.: Pokud přijede 48 lidí a budu mít tedy 12 skupin, tak 12 hráčů budou jedničky, 12 dvojky atd... dle jejich nasazení
        jednicky = serazeni_hraci[0:delic]
        dvojky = serazeni_hraci[delic:(delic)*2]
        trojky = serazeni_hraci[delic*2: (delic*3)]
        ctyrky = serazeni_hraci[delic*3: (delic*4)]
        petky = serazeni_hraci[delic*4: (delic*5)]
        sestky = serazeni_hraci[delic*5: (delic*6)]

        #Každému hráči je přiděleno identifikační číslo, podle toho v jakém koši se nachází
        for hrac in jednicky:
            hrac.kos = 1
        for hrac in dvojky:
            hrac.kos = 2
        for hrac in trojky:
            hrac.kos = 3
        for hrac in ctyrky:
            hrac.kos = 4
        for hrac in petky:
            hrac.kos = 5
        for hrac in sestky:
            hrac.kos = 6

        return jednicky, dvojky, trojky, ctyrky, petky, sestky
        
class Skupina: #Třída, která reprezentuje samotnou Skupinu turnaje
    def __init__(self, id_skupiny):
        self.id_skupiny = id_skupiny #Každá skupina má svoje id(číslo)
        self.hraci = []  #a čeká na přiřazení hráčů
    def __repr__(self):
        return f"Skupina číslo {self.id_skupiny}"
    
class Turnaj: #Třída reprezentující samotný turnaj, turnaj má svoje hráče a svoje skupiny
    def __init__(self, hraci): 
        self.hraci = hraci
        self.skupiny = []
        self.pocet_skupin = 0

    def vytvor_skupiny(self, pocet_hracu): #Metoda, která vytvoří počet skupin podle zadaných vlastností
        #Základní velikost skupiny jsou 4 hráči
        zakladni_velikost = 4 

        #Pokud přijede na turnaj méně než 7 hráčů, vytvoří se jedna skupina
        if pocet_hracu < 7: 
            self.pocet_skupin = 1

        #Pokud na turnaj přijede 13 a méně hráčů, vytvoří se dvě skupiny
        elif pocet_hracu <= 13: 
            self.pocet_skupin = 2

         #Pokud na turnaj přijede více než 13 hráčů, vytvoří se počet skupin dle vlastností
        elif pocet_hracu > 13:
            pocet_skupin_zaklad = pocet_hracu//zakladni_velikost #(1)
            zbytek = pocet_hracu% zakladni_velikost
            if zbytek%3 ==0:
                self.pocet_skupin = pocet_skupin_zaklad + 1 #(2)
            else: self.pocet_skupin = pocet_skupin_zaklad 

        #Pokud přijede počet hráčů dělitelný čtyřma, vytvoří se skupin počet dle (1), pokud po dělení 4 zbydou 1 a 2 hráči
        #Vytvoří se stejně skupin a 1 nebo 2 budou pětičlenné, pokud zbytek bude 3, vytvoří se skupina navíc a jedna ze všech skupin bude tříčlenná dle (2)
        for i in range(1, self.pocet_skupin + 1):
            nova_skupina = Skupina(id_skupiny=i)
            self.skupiny.append(nova_skupina)
    
        return self.skupiny
    
    def rozlosovani_jednicek(self,jednicky): #Metoda sloužící pro vypsání jedniček do skupin dle nasazení
        for index_hrace, hrac in enumerate(jednicky, start=1):
            for skupina in self.skupiny:
                if skupina.id_skupiny == index_hrace:
                    skupina.hraci.append(hrac)
                    break
    
    def rozlosovani_skupin(self, kos): #Metoda pro rozlosování zbytku hráčů turnaje
        while len(kos) > 0: #Pokud v koši je více než 0 hráčů
            hrac = random.choice(kos) #Vybere náhodného hráče z koše

            #Kontrola toho, aby nebyli dva hráči ze stejného klubu v jedné skupině + aby do jedné skupiny nešli dva hráči ze stejného koše
            vhodne_skupiny = []
            for skupina in self.skupiny:
                stejny_klub_nalezen = False
                stejny_kos_nalezen = False
                #Pokud se najde duplicitní klub nebo identifikace koše, tak se přeruší for cyklus a prohledává se další skupina
                for hrac_ve_skupine in skupina.hraci:
                    if hrac_ve_skupine.klub == hrac.klub:
                        stejny_klub_nalezen = True
                        break
                    elif hrac_ve_skupine.kos == hrac.kos:
                        stejny_kos_nalezen = True
                        break

                if not stejny_klub_nalezen and not stejny_kos_nalezen: #Pokud vybraný hráč nemá stejný klub, přidá se daná skupina do listu vhodných skupin
                    vhodne_skupiny.append(skupina)

            if len(vhodne_skupiny)>0: #Pokud je alespoň jedna vhodná skupina pro hráče, je tento hráč do té skupiny přidán
                min_pocet = min(len(skupina.hraci) for skupina in vhodne_skupiny) #Najde nejmenší počet hráčů
                
                kandidati = []
                for skupina in vhodne_skupiny:
                    if len(skupina.hraci) == min_pocet:
                        kandidati.append(skupina)#Pokud se počet hráčů ve vhodné skupině rovná minimu, skupina se přidá do kandidátů
                cilova_skupina = random.choice(kandidati) #Náhodně vybere kandidátní skupinu
            else: #Pokud není vhodná skupina, přidá se hráč do skupiny s nejmenším počtem hráčů (nutné porušení pravidla o duplicitě klubů)
                min_pocet = min(len(s.hraci) for s in self.skupiny) #Tady se vybírá se minimum ze všech skupin
                nejlepsi_skupiny = []
                for skupina in self.skupiny:
                    if len(skupina.hraci) == min_pocet:
                        nejlepsi_skupiny.append(skupina)#Pokud se počet hráčů v obecné skupině rovná minimu, skupina se přidá do nejlepsi_skupiny
                cilova_skupina = random.choice(nejlepsi_skupiny) #Zajištění rovnoměrného rozložení hráčů, náhodně se vybere skupina z nejlepších
            
            cilova_skupina.hraci.append(hrac)
            kos.remove(hrac)
    
    def export_turnaje_do_excelu(self, vystup, skupin_na_list):
        #Inicializace Excelovského souboru
        wb = Workbook()
        ws = wb.active
        ws.title = f"Skupiny 1-{skupin_na_list}"
        
        #Zadefinování okrajů pro každou buňku
        okraj = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
        )

        #Pomocné konstanty pro pozicování
        mezera_radky = 2
        aktualni_radek = 1
        aktualni_sloupec = 1
        #Název skupiny nad tabulkou skupiny
        for index_skupiny, skupina in enumerate(self.skupiny, start=1):
            ws.cell(row=aktualni_radek, column=aktualni_sloupec, 
                    value=f"Skupina {skupina.id_skupiny}").font= Font(bold=True, size=14)
            
            #Pomocné proměnné pro data (skok o dva řádky níže od nadpisu skupiny)
            zacatek_dat_row = aktualni_radek + 2
            zacatek_dat_collumn = aktualni_sloupec

            hraci = skupina.hraci
            n = len(hraci)
            #Příkaz pro vyplnění jmen a klubu v prním sloupci
            for i, h in enumerate(hraci):
                text = f"{h.jmeno} {h.prijmeni} ({h.nasazeni})\n{h.klub}"
                bunka = ws.cell(row= zacatek_dat_row + i, column = zacatek_dat_collumn, value=text)
                bunka.alignment = Alignment(wrap_text=True, vertical="top")
                bunka.font = Font(bold=True)

            #Pomocná proměnná pro odsazení tabulky nxn hráčů
            tabulka_zacatek_sloupec = zacatek_dat_collumn + 1

            #Hlavička hráčů
            for i in range (n):
                bunka = ws.cell(row=zacatek_dat_row -1, column= tabulka_zacatek_sloupec + i,
                        value=f"  H{i+1}  ")
                bunka.font = Font(bold=True)
                bunka.alignment = Alignment(horizontal="center", vertical="center")

            #Vytvoření čtvercové tabulky s délkou strany o počtu hráčů ve skupině + X na diagonále
            for i in range(n):
                for j in range(n):
                    if i==j:
                        bunka = ws.cell(row=zacatek_dat_row + i, column= tabulka_zacatek_sloupec + j,
                                value = "X")
                        bunka.font = Font(bold=True, size = 18)
                        bunka.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        ws.cell(row=zacatek_dat_row + i, column= tabulka_zacatek_sloupec + j,
                                value = "")
                        
            #Přidání dvou sloupců "Skóre" a "Pořadí" na konec tabulky
            sloupec_skore = tabulka_zacatek_sloupec + n
            sloupec_poradi = tabulka_zacatek_sloupec + n + 1

            bunka_skore = ws.cell(row= zacatek_dat_row - 1, column= sloupec_skore,
                    value = "Skóre")
            bunka_skore.font = Font(bold=True)
            bunka_skore.alignment = Alignment(horizontal="center", vertical="center")
            
            bunka_poradi = ws.cell(row= zacatek_dat_row - 1, column= sloupec_poradi,
                    value = "Pořadí")
            bunka_poradi.font = Font(bold=True)
            bunka_poradi.alignment = Alignment(horizontal="center", vertical="center")

            #Konec tabulky (pravý dolní roh)
            posledni_radek = zacatek_dat_row + n - 1  
            posledni_sloupec = tabulka_zacatek_sloupec + n + 1

            #Okraje pro všechny buňky skupiny
            for r in range(aktualni_radek, posledni_radek + 1):
                for c in range(aktualni_sloupec, posledni_sloupec + 1):
                    bunka = ws.cell(row=r, column=c)
                    bunka.border = okraj
            
            
            #Posun na další skupinu pod tu současnou
            aktualni_radek += n + 1 + mezera_radky

            #Vytvoření nového sheetu pokud dosáhneme požadovaného počtu skupin na jednom sheetu
            if index_skupiny % skupin_na_list == 0:

            #Zároveň se nastavuje i šířka buňky pro to, aby se tam vešly všechny texty
                for sloupec in ws.columns:
                    maximalni_delka = 0
                    for bunka in sloupec:
                        if bunka.value:
                            delka_textu = len(str(bunka.value))
                            if delka_textu > maximalni_delka:
                                maximalni_delka = delka_textu
                    pismeno_sloupce = sloupec[0].column_letter
                    ws.column_dimensions[pismeno_sloupce].width = maximalni_delka + 4
                dalsi = index_skupiny + 1
                ws = wb.create_sheet(title= f"Skupiny {dalsi}-{dalsi + skupin_na_list-1}")
                aktualni_radek = 1
                aktualni_sloupec = 1

        #Fitting slov pro poslední sheet (poslední sheet není pod podmínkou if)
        for sloupec in ws.columns:
                    maximalni_delka = 0
                    for bunka in sloupec:
                        if bunka.value:
                            delka_textu = len(str(bunka.value))
                            if delka_textu > maximalni_delka:
                                maximalni_delka = delka_textu
                    pismeno_sloupce = sloupec[0].column_letter
                    ws.column_dimensions[pismeno_sloupce].width = maximalni_delka + 4
        #Příkaz pro vytvoření Excelu            
        wb.save(vystup)
        return print(f"Turnaj rozlosován! Vygenerovaný Excel je ve stejné složce, jako tento kód")
